#!/usr/bin/env python3
"""
format_model.py — Passe de formatage unique pour pernod_ricard_valuation_v1.xlsx
Applique les conventions de la bible (couleurs semantiques, formats, en-tetes) aux 11 onglets.
- Couleurs : bleu=input, noir=formule meme-onglet, vert=lien inter-onglets.
- Idempotent. Sauvegarde horodatee avant ecriture.
- N'ecrit AUCUNE formule ni valeur : purement cosmetique (couleur police, gras, number_format).
Lancer depuis 000_script/ :  python3 format_model.py
"""
import os, shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.formula import ArrayFormula

# chemin relatif depuis 000_script/ vers le fichier dans 001_pernod_ricard/
PATH = os.path.join(os.path.dirname(__file__), "..", "001_pernod_ricard",
                    "pernod_ricard_valuation_v1.xlsx")

# --- conventions ---
BLUE, BLACK, GREEN = "FF0000FF", "FF000000", "FF008000"
PRESERVE = {"FFFFFFFF", "FF808080", "00FFFFFF"}      # headers blancs / metadonnees grises : on n'y touche pas

EUR  = "#,##0;(#,##0)"     # €m
PS   = "#,##0.00"          # par action
PCT  = "0.0%"              # pourcentage (modele, decimales)
SH   = "#,##0.000"         # actions (m)
MULT = '#,##0.0"x"'        # multiple
NPCT = '0.00"%"'           # Data : nombre-pourcent (3.74 -> 3.74%)
DATE = "yyyy-mm-dd"

# title cell + lignes d'en-tete a mettre en gras, par onglet
HEADERS = {
    "Cover": (None, []), "Data": (None, [1]), "Assumptions": ("B1", [3]),
    "Income_Statement": ("B1", [3]), "DCF": ("B1", [3]), "WACC": ("B1", [3]),
    "Bridge": ("B1", [3]), "Sensitivity": ("B1", [3, 13]), "Trading_Comps": ("B1", [3]),
    "Precedent_Transactions": ("B1", [3]), "Football_Field": ("B1", [2]),
}
TITLE_CELL = {"Cover": "B1"}  # Cover : titre en B1, pas d'en-tete tabulaire

# number_format par onglet (cellules data uniquement ; les onglets deja conformes — IS, DCF,
# Assumptions, Bridge — ne sont pas listes et ne sont donc pas touches en format)
FORMATS = {
    "Cover": [("B6",DATE),("B8",PS),("B9",PS),("B10",PCT),("B12",PS),("B13",PS),("B14",PS),
              ("C12:C14",PCT),("B15",PCT),("B16",PS),("B18:B22",PS),("C22",PS),("B23",PS),
              ("B25:B27",EUR),("B28",SH)],
    "Data": [("B2",PS),("B3",SH),("B6",PS),("B7",PS),("B8","0.00"),("B9",PS),("B10",EUR),
             ("B11",NPCT),("B13",NPCT),("B14",NPCT)],
    "WACC": [("B5",PCT),("B6",PCT),("B7","0.00"),("B8",PCT),("B10",PCT),("B11",PCT),("B12",PCT),
             ("B14",PS),("B15",SH),("B16",EUR),("B17",EUR),("B18",EUR),("B19",PCT),("B20",PCT),("B21",PCT)],
    "Sensitivity": [("C4:G4",PCT),("B5:B11",PCT),("C5:G11",PS),
                    ("C14:G14",MULT),("B15:B21",PCT),("C15:G21",PS)],
    "Trading_Comps": [("B4:B10",PS),("C4:C10",SH),("D4:F10",EUR),("G4:I10",EUR),("J4:L10",MULT),
                      ("J13:L14",MULT),("J15:L16",MULT),("B19",EUR),("B20",MULT),("B21",EUR),
                      ("B22:B24",EUR),("B25:B27",PS)],
    "Precedent_Transactions": [("D4:E10",EUR),("F4:F10",MULT),("B12:B13",MULT),("B15",EUR),
                      ("B16",MULT),("B17",EUR),("B18:B20",EUR),("B21",PS)],
    "Football_Field": [("B3:C8",PS)],
}

def is_formula(c):
    return c.data_type == "f" or isinstance(c.value, ArrayFormula) or \
           (isinstance(c.value, str) and c.value.startswith("="))

def ftext(c):
    if isinstance(c.value, ArrayFormula): return c.value.text or ""
    return c.value if isinstance(c.value, str) else ""

def set_font(c, color=None, bold=None):
    f = c.font
    c.font = Font(name=f.name, size=f.size, italic=f.italic, underline=f.underline,
                  strike=f.strike, vertAlign=f.vertAlign,
                  bold=f.bold if bold is None else bold,
                  color=f.color if color is None else color)

def each(ws, rng):
    obj = ws[rng]
    if isinstance(obj, tuple):
        for row in obj:
            for c in (row if isinstance(row, tuple) else (row,)):
                yield c
    else:
        yield obj

def main():
    path = os.path.abspath(PATH)
    if not os.path.exists(path):
        raise SystemExit(f"Fichier introuvable : {path}")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = f"{os.path.splitext(path)[0]}_backup_{ts}.xlsx"
    shutil.copy(path, backup)
    print(f"Backup -> {backup}")

    wb = load_workbook(path)

    # 1) number formats
    for sheet, rules in FORMATS.items():
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        for rng, fmt in rules:
            for c in each(ws, rng):
                if not isinstance(c, MergedCell):
                    c.number_format = fmt

    # 2) couleurs semantiques (tous les onglets, idempotent)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c, MergedCell) or c.value is None:
                    continue
                col = c.font.color
                rgb = col.rgb if (col is not None and col.type == "rgb") else None
                if rgb in PRESERVE:
                    continue
                if is_formula(c):
                    set_font(c, color=GREEN if "!" in ftext(c) else BLACK)
                elif isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                    set_font(c, color=BLUE)
                else:
                    set_font(c, color=BLACK)

    # 3) gras titres + en-tetes
    for sheet, (title, hrows) in HEADERS.items():
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        tcell = TITLE_CELL.get(sheet, title)
        if tcell and ws[tcell].value is not None:
            set_font(ws[tcell], bold=True)
        for r in hrows:
            for c in ws[r]:
                if c.value is not None and not isinstance(c, MergedCell):
                    set_font(c, bold=True)

    wb.save(path)
    print(f"Formatage applique et sauvegarde -> {path}")

if __name__ == "__main__":
    main()